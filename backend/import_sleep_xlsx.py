import argparse
import os
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Iterable, List, Optional, Tuple

from zoneinfo import ZoneInfo


DEFAULT_USER_NAME = "Паша"


MOSCOW_TZ = ZoneInfo("Europe/Moscow")
UTC_TZ = ZoneInfo("UTC")


TIME_RANGE_RE = re.compile(r"(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})")


@dataclass(frozen=True)
class SleepInterval:
    start_local: datetime
    end_local: datetime
    kind: str  # day|night


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Импорт сна из XLSX (таблица с днями по колонкам).")
    p.add_argument("--user", default=DEFAULT_USER_NAME, help="Имя пользователя, в которого импортируем сон.")
    p.add_argument("--xlsx", required=True, help="Путь к XLSX-файлу.")
    p.add_argument("--dry-run", action="store_true", help="Проверка без записи в БД.")
    p.add_argument(
        "--wipe-all-sleep-logs",
        action="store_true",
        help="ОПАСНО: удалить ВСЕ sleep-логи ВСЕХ пользователей перед импортом (требует --yes, если не dry-run).",
    )
    p.add_argument(
        "--wipe-user-sleep-logs",
        action="store_true",
        help="Удалить sleep-логи выбранного пользователя перед импортом.",
    )
    p.add_argument(
        "--yes",
        action="store_true",
        help="Подтверждение опасных операций (нужно для --wipe-all-sleep-logs в режиме записи).",
    )
    return p.parse_args()


def _parse_sheet_dates(ws) -> dict:
    # Ищем строку, где в A есть "Дата".
    date_row = None
    for r in range(1, 8):
        v = ws.cell(r, 1).value
        if str(v).strip() == "Дата":
            date_row = r
            break
    if not date_row:
        raise ValueError("Не найдена строка 'Дата' в листе")

    dates = {}
    for c in range(2, 500):
        v = ws.cell(date_row, c).value
        if v in (None, ""):
            continue
        if isinstance(v, datetime):
            d = v.date()
        else:
            s = str(v).strip()
            try:
                d = datetime.strptime(s, "%d.%m.%Y").date()
            except Exception:
                continue
        dates[c] = d
    if not dates:
        raise ValueError("Не найдены даты в строке 'Дата'")
    return {"date_row": date_row, "dates": dates}


def _parse_sleep_intervals_from_cell(day_date, raw) -> List[SleepInterval]:
    if raw in (None, ""):
        return []
    text = str(raw)
    if "Ночной сон" in text:
        kind = "night"
    elif "Дневной сон" in text:
        kind = "day"
    else:
        return []  # бодрствование / промежутки без меток — не импортируем

    out: List[SleepInterval] = []
    for m in TIME_RANGE_RE.finditer(text):
        a, b = m.group(1), m.group(2)
        start_local = datetime(day_date.year, day_date.month, day_date.day, int(a[:2]), int(a[3:]), tzinfo=MOSCOW_TZ)
        end_local = datetime(day_date.year, day_date.month, day_date.day, int(b[:2]), int(b[3:]), tzinfo=MOSCOW_TZ)
        if end_local <= start_local:
            end_local = end_local + timedelta(days=1)
        out.append(SleepInterval(start_local=start_local, end_local=end_local, kind=kind))
    return out


def _iter_sleep_intervals(ws) -> Iterable[SleepInterval]:
    meta = _parse_sheet_dates(ws)
    date_row = meta["date_row"]
    dates = meta["dates"]

    # Сканируем тело листа ниже строки с датами.
    for c, day_date in sorted(dates.items(), key=lambda x: x[1]):
        for r in range(date_row + 1, date_row + 120):
            raw = ws.cell(r, c).value
            for it in _parse_sleep_intervals_from_cell(day_date, raw):
                yield it


def main() -> None:
    args = parse_args()
    print(f"=== ИМПОРТ СНА ({args.user}) ===")
    print(f"XLSX: {args.xlsx}")
    print(f"Режим: {'DRY-RUN' if args.dry_run else 'WRITE'}")

    if not os.path.exists(args.xlsx):
        print(f"❌ Файл не найден: {args.xlsx}")
        return

    if args.wipe_all_sleep_logs and args.wipe_user_sleep_logs:
        print("❌ Нельзя одновременно использовать --wipe-all-sleep-logs и --wipe-user-sleep-logs")
        return

    if args.wipe_all_sleep_logs and (not args.dry_run) and (not args.yes):
        print("❌ Для удаления ВСЕХ sleep-логов нужен флаг подтверждения --yes")
        print("   Пример: python import_sleep_xlsx.py --xlsx file.xlsx --wipe-all-sleep-logs --yes")
        return

    from openpyxl import load_workbook
    from app.database import SessionLocal, engine, Base
    from app import models

    Base.metadata.create_all(bind=engine)

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    if not wb.sheetnames:
        print("❌ В книге нет листов")
        return
    ws = wb[wb.sheetnames[0]]
    print(f"Лист: {ws.title}")

    db = SessionLocal()
    try:
        user = db.query(models.User).filter(models.User.name == args.user).first()
        if not user:
            print(f"❌ Пользователь '{args.user}' не найден")
            return

        if args.wipe_all_sleep_logs:
            total = db.query(models.BabyLog).filter(models.BabyLog.event_type == "sleep").count()
            print(f"🧨 Удаление ВСЕХ sleep-логов в БД: {total}")
            if not args.dry_run:
                db.query(models.BabyLog).filter(models.BabyLog.event_type == "sleep").delete(synchronize_session=False)
                db.commit()

        if args.wipe_user_sleep_logs:
            n = (
                db.query(models.BabyLog)
                .filter(models.BabyLog.event_type == "sleep", models.BabyLog.user_id == user.id)
                .count()
            )
            print(f"🧹 Удаление sleep-логов пользователя: {n}")
            if not args.dry_run:
                (
                    db.query(models.BabyLog)
                    .filter(models.BabyLog.event_type == "sleep", models.BabyLog.user_id == user.id)
                    .delete(synchronize_session=False)
                )
                db.commit()

        intervals = list(_iter_sleep_intervals(ws))
        if not intervals:
            print("⚠️ Не найдено интервалов сна (Дневной/Ночной сон).")
            return

        created = 0
        for it in intervals:
            start_utc = it.start_local.astimezone(UTC_TZ).replace(tzinfo=None)
            end_utc = it.end_local.astimezone(UTC_TZ).replace(tzinfo=None)
            duration = int((it.end_local - it.start_local).total_seconds() // 60)
            if duration <= 0:
                continue
            note = f"import:xlsx; kind={it.kind}; tz=Europe/Moscow"
            row = models.BabyLog(
                user_id=user.id,
                creator_name_snapshot=args.user,
                event_type="sleep",
                start_time=start_utc,
                end_time=end_utc,
                duration_minutes=duration,
                note=note,
            )
            db.add(row)
            created += 1

        print(f"✅ Интервалов сна к добавлению: {created}")
        if args.dry_run:
            db.rollback()
            print("DRY-RUN: изменения не записаны.")
        else:
            db.commit()
            print("WRITE: импорт завершён.")

    finally:
        db.close()


if __name__ == "__main__":
    main()

