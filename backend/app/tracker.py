from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_
from datetime import datetime, timezone, timedelta, time
from typing import List, Optional, Dict, Any
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from fastapi.responses import StreamingResponse

from .database import get_db
from .models import BabyLog, User
from .schemas import BabyLogCreate, BabyLogOut, BabyLogUpdate, BabyLogPageOut
from .auth import get_current_user

from fastapi import Request
from .rate_limit import limiter

from zoneinfo import ZoneInfo

router = APIRouter()

def get_utc_now():
    return datetime.now(timezone.utc)


MOSCOW_TZ = ZoneInfo("Europe/Moscow")

@router.get("/stats", response_model=Dict[str, Any])
def get_tracker_stats(
    days: int = 7,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    now = get_utc_now()
    # Логи пишем/храним в UTC, но "день" и период считаем по Москве.
    now_local = now.astimezone(MOSCOW_TZ)
    days = max(1, min(int(days or 7), 365))
    start_local_date = (now_local - timedelta(days=days - 1)).date()
    start_local_dt = datetime.combine(start_local_date, time.min, tzinfo=MOSCOW_TZ)
    start_date = start_local_dt.astimezone(timezone.utc).replace(tzinfo=None)

    logs = db.query(BabyLog).filter(
        BabyLog.user_id == current_user.id,
        BabyLog.event_type == "sleep",
        BabyLog.start_time >= start_date
    ).order_by(BabyLog.start_time.asc()).all()

    daily_stats = []
    stats_map = {}
    
    for i in range(days):
        d = (now_local - timedelta(days=i)).date()
        key = d.isoformat()
        stats_map[key] = {"date": key, "sleep_minutes": 0, "sessions_count": 0}

    total_sleep_minutes = 0
    total_sessions = 0

    for log in logs:
        if not log.end_time:
            continue 
        
        duration_min = log.duration_minutes
        if duration_min is None:
            delta = log.end_time - log.start_time
            duration_min = int(delta.total_seconds() / 60)
        
        # ВАЖНО: ночной сон не режем по полуночи — целиком относим в день старта (по Москве).
        st = log.start_time
        if st.tzinfo is None:
            st = st.replace(tzinfo=timezone.utc)
        day_key = st.astimezone(MOSCOW_TZ).date().isoformat()
        
        if day_key in stats_map:
            stats_map[day_key]["sleep_minutes"] += duration_min
            stats_map[day_key]["sessions_count"] += 1
            
        total_sleep_minutes += duration_min
        total_sessions += 1

    daily_stats = sorted(stats_map.values(), key=lambda x: x["date"])

    return {
        "period_days": days,
        "total_sleep_minutes": total_sleep_minutes,
        "total_sessions": total_sessions,
        "average_sleep_minutes": int(total_sleep_minutes / total_sessions) if total_sessions > 0 else 0,
        "daily_breakdown": daily_stats
    }

@router.get("/status", response_model=Dict[str, Any])
def get_tracker_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    active_sleep = db.query(BabyLog).filter(
        BabyLog.user_id == current_user.id,
        BabyLog.event_type == "sleep",
        BabyLog.end_time == None
    ).first()

    last_sleep = db.query(BabyLog).filter(
        BabyLog.user_id == current_user.id,
        BabyLog.event_type == "sleep",
        BabyLog.end_time != None
    ).order_by(BabyLog.end_time.desc()).first()

    last_wake_time = last_sleep.end_time if last_sleep else None

    if active_sleep:
        return {
            "is_sleeping": True,
            "current_sleep_id": active_sleep.id,
            "current_sleep_start": active_sleep.start_time.isoformat(),
            "last_wake_up": None
        }
    else:
        return {
            "is_sleeping": False,
            "current_sleep_id": None,
            "current_sleep_start": None,
            "last_wake_up": last_wake_time.isoformat() if last_wake_time else None
        }

def _parse_log_cursor_time(value: Optional[str], field: str) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid {field} format")


@router.get("/logs", response_model=BabyLogPageOut)
def get_logs(
    limit: int = Query(100, ge=1, le=500),
    after_start_time: Optional[str] = Query(
        None, description="Курсор: start_time + id (новые сначала)"
    ),
    after_id: Optional[int] = None,
    event_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if (after_start_time is None) != (after_id is None):
        raise HTTPException(
            status_code=400,
            detail="Укажите оба параметра: after_start_time и after_id, либо ни одного",
        )

    q = db.query(BabyLog).filter(BabyLog.user_id == current_user.id)
    if event_type:
        q = q.filter(BabyLog.event_type == event_type)

    total = q.count()

    if after_start_time is not None and after_id is not None:
        c_t = _parse_log_cursor_time(after_start_time, "after_start_time")
        if c_t is None:
            raise HTTPException(status_code=400, detail="Invalid after_start_time")
        q = q.filter(
            or_(
                BabyLog.start_time < c_t,
                and_(BabyLog.start_time == c_t, BabyLog.id < after_id),
            )
        )

    take = min(limit + 1, 501)
    rows = q.order_by(BabyLog.start_time.desc(), BabyLog.id.desc()).limit(take).all()
    has_more = len(rows) > limit
    if has_more:
        rows = rows[:limit]

    result: List[BabyLogOut] = []
    for log in rows:
        is_active = log.end_time is None
        log_dict = BabyLogOut.from_orm(log)
        log_dict.creator_name = log.creator_name_snapshot if log.creator_name_snapshot else "Пользователь (удален)"
        log_dict.is_active = is_active
        result.append(log_dict)

    return BabyLogPageOut(items=result, has_more=has_more, total=total)

@router.post("/logs", response_model=BabyLogOut)
def create_log(
    log_in: BabyLogCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if log_in.event_type == "sleep":
        existing_active = db.query(BabyLog).filter(
            BabyLog.user_id == current_user.id,
            BabyLog.event_type == "sleep",
            BabyLog.end_time == None
        ).first()
        if existing_active:
            raise HTTPException(status_code=400, detail="Уже идет активный сон. Завершите его сначала.")

    duration = 0
    if log_in.end_time and log_in.start_time:
        delta = log_in.end_time - log_in.start_time
        duration = int(delta.total_seconds() / 60)
    
    db_log = BabyLog(
        user_id=current_user.id,
        event_type=log_in.event_type,
        start_time=log_in.start_time,
        end_time=log_in.end_time,
        duration_minutes=duration,
        note=log_in.note,
        creator_name_snapshot=current_user.name
    )
    
    db.add(db_log)
    db.commit()
    db.refresh(db_log)
    
    res = BabyLogOut.from_orm(db_log)
    res.creator_name = db_log.creator_name_snapshot
    res.is_active = (db_log.end_time is None)
    return res

@router.post("/logs/quick-feed", response_model=BabyLogOut)
def quick_feed(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    now = get_utc_now()
    db_log = BabyLog(
        user_id=current_user.id,
        event_type="feed",
        start_time=now,
        end_time=now,
        duration_minutes=0,
        note="Быстрое добавление",
        creator_name_snapshot=current_user.name
    )
    
    db.add(db_log)
    db.commit()
    db.refresh(db_log)
    
    res = BabyLogOut.from_orm(db_log)
    res.creator_name = db_log.creator_name_snapshot
    res.is_active = False
    return res

@router.put("/logs/{log_id}", response_model=BabyLogOut)
def update_log(
    log_id: int,
    log_in: BabyLogUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    db_log = db.query(BabyLog).filter(BabyLog.id == log_id, BabyLog.user_id == current_user.id).first()
    if not db_log:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    
    update_data = log_in.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_log, field, value)
    
    if db_log.start_time and db_log.end_time:
        start = db_log.start_time
        end = db_log.end_time
        
        if start.tzinfo is None:
            start = start.replace(tzinfo=timezone.utc)
        if end.tzinfo is None:
            end = end.replace(tzinfo=timezone.utc)
            
        delta = end - start
        db_log.duration_minutes = int(delta.total_seconds() / 60)
    elif db_log.end_time is None:
        db_log.duration_minutes = 0        
        
    db.commit()
    db.refresh(db_log)
    
    res = BabyLogOut.from_orm(db_log)
    res.creator_name = db_log.creator_name_snapshot
    res.is_active = (db_log.end_time is None)
    return res

@router.delete("/logs/{log_id}")
def delete_log(
    log_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    db_log = db.query(BabyLog).filter(BabyLog.id == log_id, BabyLog.user_id == current_user.id).first()
    if not db_log:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    
    db.delete(db_log)
    db.commit()
    return {"status": "deleted"}

# ==========================================
# ЭКСПОРТ В EXCEL (ОБНОВЛЕННЫЙ)
# ==========================================

@router.get("/export/excel")
@limiter.limit("3/minute")
def export_tracker_excel(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    _ = current_user
    query = db.query(BabyLog)
    
    if start_date:
        try:
            sd = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.filter(BabyLog.start_time >= sd)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date format")
            
    if end_date:
        try:
            ed = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.filter(BabyLog.start_time <= ed)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid end_date format")
            
    sessions = query.order_by(BabyLog.start_time.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sleep History"
    
    # Заголовки
    headers = ["ID", "Event Type", "Start Time", "End Time", "Duration (Min)", "Duration (Hours)", "Note", "Creator"]
    ws.append(headers)
    
    # Стили
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
    # Данные
    for s in sessions:
        duration_min = s.duration_minutes
        if duration_min is None and s.end_time and s.start_time:
            delta = s.end_time - s.start_time
            duration_min = int(delta.total_seconds() / 60)
            
        duration_hours = round(duration_min / 60.0, 2) if duration_min else 0
        
        start_val = s.start_time.strftime("%Y-%m-%d %H:%M") if isinstance(s.start_time, datetime) else str(s.start_time)
        end_val = s.end_time.strftime("%Y-%m-%d %H:%M") if isinstance(s.end_time, datetime) else str(s.end_time)
        
        row = [
            s.id,
            s.event_type,
            start_val,
            end_val,
            duration_min,
            duration_hours,
            s.note or "",
            s.creator_name_snapshot or "Unknown"
        ]
        ws.append(row)
    
    # Авто-ширина
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)
        
    # Заморозка шапки
    ws.freeze_panes = "A2"
    
    # Умная таблица
    max_row = ws.max_row
    max_col = ws.max_column
    tab_ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
    
    table = Table(displayName="SleepTable", ref=tab_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    
    ws.add_table(table)
    
    # Сохранение
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"sleep_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )