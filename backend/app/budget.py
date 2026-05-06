from fastapi import APIRouter, Depends, HTTPException, status, Response, Query, Body
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, and_, func, select
from typing import List, Optional
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from fastapi.responses import StreamingResponse

from . import models, schemas
from .auth import get_current_user
from .database import get_db
from .models import User

from fastapi import Request
from .rate_limit import limiter

router = APIRouter()

# ==========================================
# КАТЕГОРИИ (ГРУППЫ)
# ==========================================

@router.get("/groups", response_model=List[schemas.CategoryGroupOut])
def get_groups(db: Session = Depends(get_db)):
    """Получить все категории (группы) вместе с подкатегориями."""
    print("📡 [BACKEND] Запрос /groups получен")
    
    groups = db.query(models.CategoryGroup).options(
        joinedload(models.CategoryGroup.subcategories)
    ).all()
    
    result = []
    for g in groups:
        # Логируем состояние группы
        print(f"   📂 Группа: {g.name} (ID: {g.id}) | Скрыта: {g.is_hidden}")
        
        group_dict = schemas.CategoryGroupOut.from_orm(g)
        
        for sub in group_dict.subcategories:
            # Находим оригинальный объект из БД для проверки флага
            # (на случай если ORM кэширует что-то не то, хотя не должно)
            db_sub = next((s for s in g.subcategories if s.id == sub.id), None)
            real_is_hidden = db_sub.is_hidden if db_sub else sub.is_hidden
            
            # Логируем состояние подкатегории
            print(f"      ↳ Подкат: {sub.name} (ID: {sub.id}) | Скрыта: {real_is_hidden}")
            
            sub.group_name = g.name
        
        result.append(group_dict)
    
    print(f"✅ [BACKEND] Отправлено {len(result)} групп")
    return result
    
@router.post("/groups", response_model=schemas.CategoryGroupOut, status_code=201)
def create_group(group_in: schemas.CategoryGroupCreate, db: Session = Depends(get_db)):
    db_group = models.CategoryGroup(**group_in.dict())
    db.add(db_group)
    db.commit()
    db.refresh(db_group)
    return db_group

@router.put("/groups/{group_id}", response_model=schemas.CategoryGroupOut)
def update_group(group_id: int, group_in: schemas.CategoryGroupCreate, db: Session = Depends(get_db)):
    obj = db.query(models.CategoryGroup).filter(models.CategoryGroup.id == group_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Категория не найдена")
    
    for field, value in group_in.dict().items():
        setattr(obj, field, value)
    
    db.commit()
    db.refresh(obj)
    return obj

@router.delete("/groups/{group_id}")
def delete_group(
    group_id: int, 
    db: Session = Depends(get_db),
    force: bool = Query(False, description="Если True, удаляет группу и все подкатегории. Иначе - скрывает.")
):
    obj = db.query(models.CategoryGroup).filter(models.CategoryGroup.id == group_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Категория не найдена")
    
    if force:
        # Жесткое удаление: удалит группу и все подкатегории (cascade).
        # Транзакции останутся, но связь с подкатегориями будет разорвана (если настроено ON DELETE SET NULL) 
        # или удалена (если CASCADE). В моделях у нас Cascade на подкатегории, а у подкатегорий связь с транзакциями.
        # Внимание: Если подкатегории удалятся, транзакции могут потерять категорию.
        # В нашей модели Transaction.category_id NOT NULL, поэтому нужно быть осторожным.
        # Лучшая стратегия при force: сначала скрыть, либо запретить удаление если есть транзакции.
        # Для простоты пока просто удаляем группу и подкатегории. Транзакции останутся с ID несуществующей категории? 
        # Нет, FK защитит. 
        # РЕШЕНИЕ: При force мы просто скрываем всё глубоко, или требуем ручного удаления транзакций пользователем.
        # Пока реализуем мягкое удаление даже для force, но с флагом permanent_hidden? 
        # Нет, давай сделаем так: Force удаляет только если нет транзакций в подкатегориях.
        
        has_tx = False
        for sub in obj.subcategories:
            count = db.query(models.Transaction).filter(models.Transaction.category_id == sub.id).count()
            if count > 0:
                has_tx = True
                break
        
        if has_tx:
            raise HTTPException(status_code=400, detail="Невозможно удалить: в этой категории или подкатегориях есть транзакции. Сначала удалите их.")
            
        db.delete(obj)
        db.commit()
        return {"status": "deleted"}
    else:
        # Мягкое удаление (скрытие)
        obj.is_hidden = True
        # Скрываем все подкатегории тоже
        for sub in obj.subcategories:
            sub.is_hidden = True
        db.commit()
        return {"status": "hidden"}

@router.post("/groups/{group_id}/unhide")
def unhide_group(group_id: int, db: Session = Depends(get_db)):
    obj = db.query(models.CategoryGroup).filter(models.CategoryGroup.id == group_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Группа не найдена")
    
    obj.is_hidden = False
    db.commit()
    db.refresh(obj)
    return {"status": "unhidden"}

# ==========================================
# ПОДКАТЕГОРИИ
# ==========================================

@router.post("/subcategories", response_model=schemas.CategoryOut, status_code=201)
def create_subcategory(sub_in: schemas.CategoryCreate, db: Session = Depends(get_db)):
    # Проверка существования группы
    group = db.get(models.CategoryGroup, sub_in.group_id)
    if not group:
        raise HTTPException(status_code=400, detail="Родительская категория не найдена")
    
    db_sub = models.Category(**sub_in.dict())
    db.add(db_sub)
    db.commit()
    db.refresh(db_sub)
    
    res = schemas.CategoryOut.from_orm(db_sub)
    res.group_name = group.name
    return res

@router.put("/subcategories/{sub_id}", response_model=schemas.CategoryOut)
def update_subcategory(sub_id: int, sub_in: schemas.CategoryCreate, db: Session = Depends(get_db)):
    obj = db.query(models.Category).filter(models.Category.id == sub_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Подкатегория не найдена")
    
    # Если меняем группу, проверяем её
    if sub_in.group_id != obj.group_id:
        group = db.get(models.CategoryGroup, sub_in.group_id)
        if not group:
            raise HTTPException(status_code=400, detail="Новая родительская категория не найдена")
    
    for field, value in sub_in.dict().items():
        setattr(obj, field, value)
        
    db.commit()
    db.refresh(obj)
    
    res = schemas.CategoryOut.from_orm(obj)
    if obj.group:
        res.group_name = obj.group.name
    return res

@router.delete("/subcategories/{sub_id}")
def delete_subcategory(
    sub_id: int,
    db: Session = Depends(get_db),
    force: bool = Query(False, description="Если True, удаляет подкатегорию. Иначе - скрывает.")
):
    obj = db.query(models.Category).filter(models.Category.id == sub_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Подкатегория не найдена")
    
    if force:
        # Проверка на наличие транзакций
        count = db.query(models.Transaction).filter(models.Transaction.category_id == sub_id).count()
        if count > 0:
            raise HTTPException(status_code=400, detail=f"Невозможно удалить: имеется {count} транзакций. Удалите их сначала.")
        
        db.delete(obj)
        db.commit()
        return {"status": "deleted"}
    else:
        # Скрытие
        obj.is_hidden = False if obj.is_hidden else True # Toggle или просто True? Давайте просто True (скрыть)
        obj.is_hidden = True
        db.commit()
        return {"status": "hidden"}

@router.post("/subcategories/{sub_id}/unhide")
def unhide_subcategory(sub_id: int, db: Session = Depends(get_db)):
    obj = db.query(models.Category).filter(models.Category.id == sub_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Подкатегория не найдена")
    
    # Проверка: скрыт ли родитель? (Опционально, по вашему ТЗ)
    # if obj.group and obj.group.is_hidden:
    #     raise HTTPException(status_code=400, detail="Сначала восстановите родительскую категорию")
        
    obj.is_hidden = False
    db.commit()
    db.refresh(obj)
    return {"status": "unhidden"}

# ==========================================
# ТРАНЗАКЦИИ (Обновлено для новых путей)
# ==========================================

def make_tx_resp(t, current_balance=None):
    if not t:
        return None
    
    # Формирование пути: "Категория / Подкатегория"
    path = "Нет категории"
    if t.category:
        sub_name = t.category.name
        group_name = t.category.group.name if t.category.group else "Без категории"
        path = f"{group_name} / {sub_name}"
    
    date_str = t.date.strftime("%Y-%m-%dT%H:%M:%S") if isinstance(t.date, datetime) else str(t.date)
    
    creator_display_name = "Неизвестно"
    if t.creator_name_snapshot:
        creator_display_name = t.creator_name_snapshot
    elif t.creator:
        creator_display_name = t.creator.name
    
    resp = schemas.TransactionOut(
        id=t.id, 
        amount=t.amount, 
        transaction_type=t.transaction_type,
        category_id=t.category_id, 
        description=t.description,
        date=date_str,
        creator_name=creator_display_name,
        full_category_path=path
    )
    if current_balance is not None:
        resp.balance = current_balance
    return resp

def _parse_query_datetime(value: Optional[str], field: str) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid {field} format")


@router.get(
    "/transactions",
    response_model=schemas.TransactionPageOut,
    dependencies=[Depends(get_current_user)],
)
def get_transactions(
    response: Response,
    db: Session = Depends(get_db),
    limit: int = Query(100, ge=1, le=500, description="Размер страницы"),
    after_date: Optional[str] = Query(
        None, description="Курсор: дата последней выданной строки (с newest-first)"
    ),
    after_id: Optional[int] = Query(None, description="Курсор: id (вместе с after_date)"),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    category_id: Optional[int] = None,
    group_id: Optional[int] = None,
    created_by_user_id: Optional[int] = Query(
        None,
        description="Опциональный фильтр по автору; по умолчанию видны все операции",
    ),
):
    """
    Список операций: лента для всех; фильтры (даты, категории) — в query, как «что подглядываем».
    Баланс на карточке — глобальный накопительный, как в старой выдаче: SUM(amount) по всем операциям
    по мере (date, id) без сужения под фильтр.
    """
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"

    if (after_date is None) != (after_id is None):
        raise HTTPException(
            status_code=400,
            detail="Укажите оба параметра: after_date и after_id, либо ни одного",
        )

    d_from = _parse_query_datetime(date_from, "date_from")
    d_to = _parse_query_datetime(date_to, "date_to")

    t = models.Transaction.__table__
    # Глобальный бегущий total по всему движению; фильтры — только внешним SELECT
    run_bal = (
        func.sum(t.c.amount)
        .over(order_by=(t.c.date.asc(), t.c.id.asc()))
        .label("running_balance")
    )
    inner = select(t, run_bal).select_from(t)
    sub = inner.subquery("tx_win")

    sub_ids: Optional[List[int]] = None
    if group_id is not None:
        sub_ids = [r[0] for r in db.query(models.Category.id).filter(models.Category.group_id == group_id).all()]
        if not sub_ids:
            return schemas.TransactionPageOut(items=[], has_more=False, total=0)

    q_count = db.query(models.Transaction)
    if created_by_user_id is not None:
        q_count = q_count.filter(models.Transaction.created_by_user_id == created_by_user_id)
    if d_from is not None:
        q_count = q_count.filter(models.Transaction.date >= d_from)
    if d_to is not None:
        q_count = q_count.filter(models.Transaction.date <= d_to)
    if category_id is not None:
        q_count = q_count.filter(models.Transaction.category_id == category_id)
    if sub_ids is not None:
        q_count = q_count.filter(models.Transaction.category_id.in_(sub_ids))
    total = q_count.count()

    outer = select(sub)
    if created_by_user_id is not None:
        outer = outer.where(sub.c.created_by_user_id == created_by_user_id)
    if d_from is not None:
        outer = outer.where(sub.c.date >= d_from)
    if d_to is not None:
        outer = outer.where(sub.c.date <= d_to)
    if category_id is not None:
        outer = outer.where(sub.c.category_id == category_id)
    if sub_ids is not None:
        outer = outer.where(sub.c.category_id.in_(sub_ids))
    if after_date is not None and after_id is not None:
        c_dt = _parse_query_datetime(after_date, "after_date")
        if c_dt is None:
            raise HTTPException(status_code=400, detail="Invalid after_date")
        outer = outer.where(
            or_(
                sub.c.date < c_dt,
                and_(sub.c.date == c_dt, sub.c.id < after_id),
            )
        )
    take = min(limit + 1, 501)
    outer = outer.order_by(sub.c.date.desc(), sub.c.id.desc()).limit(take)

    result_rows = list(db.execute(outer).mappings().all())
    has_more = len(result_rows) > limit
    if has_more:
        result_rows = result_rows[:limit]

    if not result_rows:
        return schemas.TransactionPageOut(items=[], has_more=has_more, total=total)

    page_ids = [int(m["id"]) for m in result_rows]
    id_to_balance = {int(m["id"]): float(m["running_balance"] or 0.0) for m in result_rows}

    orm_txs = (
        db.query(models.Transaction)
        .options(
            joinedload(models.Transaction.category).joinedload(models.Category.group)
        )
        .filter(models.Transaction.id.in_(page_ids))
        .all()
    )
    by_id = {x.id: x for x in orm_txs}
    items = []
    for tid in page_ids:
        t_obj = by_id.get(tid)
        if t_obj is None:
            continue
        items.append(make_tx_resp(t_obj, id_to_balance.get(tid, 0.0)))
    return schemas.TransactionPageOut(items=items, has_more=has_more, total=total)

@router.post("/transactions", response_model=schemas.TransactionOut, status_code=201)
def create_transaction(
    tx: schemas.TransactionCreate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Если категория не передана, используем заглушку
    category_id = tx.category_id
    
    if category_id is None:
        default_group = db.query(models.CategoryGroup).filter(models.CategoryGroup.name == "Без категории").first()
        if default_group:
            default_sub = db.query(models.Category).filter(models.Category.group_id == default_group.id).first()
            if default_sub:
                category_id = default_sub.id
        if category_id is None:
            raise HTTPException(status_code=500, detail="Системная ошибка: не найдена категория по умолчанию")

    # Проверка существования категории (на всякий случай)
    cat = db.get(models.Category, category_id)
    if not cat:
        raise HTTPException(status_code=400, detail="Подкатегория не найдена")
    
    db_tx = models.Transaction(
        amount=tx.amount, 
        transaction_type=tx.transaction_type,
        category_id=category_id, # Используем проверенный ID
        description=tx.description,
        date=tx.date, 
        created_by_user_id=current_user.id,
        creator_name_snapshot=current_user.name
    )
    
    db.add(db_tx)
    db.commit()
    db.refresh(db_tx)
    db_tx = (
        db.query(models.Transaction)
        .options(
            joinedload(models.Transaction.category).joinedload(models.Category.group)
        )
        .filter(models.Transaction.id == db_tx.id)
        .first()
    )
    if not db_tx:
        raise HTTPException(status_code=500, detail="Ошибка чтения созданной операции")
    sum_row = (
        db.query(func.coalesce(func.sum(models.Transaction.amount), 0.0))
        .filter(
            or_(
                models.Transaction.date < db_tx.date,
                and_(models.Transaction.date == db_tx.date, models.Transaction.id <= db_tx.id),
            )
        )
        .scalar()
    )
    return make_tx_resp(db_tx, float(sum_row or 0.0))

@router.get("/transactions/{tx_id}", response_model=schemas.TransactionOut)
def get_transaction(tx_id: int, db: Session = Depends(get_db)):
    t = db.query(models.Transaction).options(
        joinedload(models.Transaction.category).joinedload(models.Category.group)
    ).filter(models.Transaction.id == tx_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Not found")
    return make_tx_resp(t)

@router.put("/transactions/{tx_id}", response_model=schemas.TransactionOut)
def update_transaction(tx_id: int, tx: schemas.TransactionUpdate, db: Session = Depends(get_db)):
    obj = db.query(models.Transaction).filter(models.Transaction.id == tx_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Not found")
    
    update_data = tx.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(obj, key, value)
    
    db.commit()
    db.refresh(obj)
    return make_tx_resp(obj)

@router.delete("/transactions/{tx_id}")
def delete_transaction(tx_id: int, db: Session = Depends(get_db)):
    obj = db.query(models.Transaction).filter(models.Transaction.id == tx_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Not found")
    
    db.delete(obj)
    db.commit()
    return {"status": "deleted"}

# ==========================================
# ЭКСПОРТ (Обновлен под новые поля)
# ==========================================

@router.get("/export/excel")
@limiter.limit("3/minute")
def export_budget_excel(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    _ = current_user
    query = db.query(models.Transaction).options(
        joinedload(models.Transaction.category).joinedload(models.Category.group),
    )

    if start_date:
        try:
            sd = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.filter(models.Transaction.date >= sd)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date format")
            
    if end_date:
        try:
            ed = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.filter(models.Transaction.date <= ed)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid end_date format")
            
    transactions = query.order_by(models.Transaction.date.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget History"
    
    headers = ["ID", "Date", "Amount", "Type", "Category", "Subcategory", "Description", "Creator"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
    for t in transactions:
        cat_name = t.category.group.name if (t.category and t.category.group) else ""
        subcat_name = t.category.name if t.category else ""
        
        date_val = t.date.strftime("%Y-%m-%d %H:%M") if isinstance(t.date, datetime) else str(t.date)
        
        row = [
            t.id,
            date_val,
            t.amount,
            t.transaction_type,
            cat_name,
            subcat_name,
            t.description or "",
            t.creator_name_snapshot or "Unknown"
        ]
        ws.append(row)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)
        
    ws.freeze_panes = "A2"
    
    max_row = ws.max_row
    max_col = ws.max_column
    tab_ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
    
    table = Table(displayName="BudgetTable", ref=tab_ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    
    ws.add_table(table)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"budget_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )